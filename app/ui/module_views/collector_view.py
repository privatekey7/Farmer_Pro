# app/ui/module_views/collector_view.py
from __future__ import annotations
from dataclasses import dataclass

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QLabel, QLineEdit, QFileDialog, QPushButton,
    QCheckBox, QDoubleSpinBox, QSpinBox, QScrollArea, QFrame,
)

from app.ui.widgets.chain_picker import ChainPickerWidget
from app.ui.widgets.drop_zone import DropZone

from app.core.models import ProxyConfig
from app.storage.parsers import parse_wallets, parse_proxies, parse_lines
from app.i18n import tr, i18n


@dataclass
class CollectorSettings:
    min_token_usd: float          # default 1.0
    min_bridge_usd: float         # default 1.0 — пропустить бридж если баланс меньше
    excluded_chains: list[str]    # DeBank keys в lowercase
    target_chains: list[str]      # DeBank keys или названия сетей
    slippage: float               # десятичная (0.005 = 0.5%)
    delay_min: int                # default 10
    delay_max: int                # default 30
    send_to_exchange: bool
    delay_after_bridge: int       # default 60


class CollectorConfigWidget(QWidget):
    """Панель настроек модуля Collector."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._wallets: list[dict] = []       # [{"raw": ..., "type": ...}]
        self._proxies: list[ProxyConfig] = []
        self._subaccounts: list[str] = []
        self._export_results: list = []
        self._export_details: dict = {}

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(8)

        scroll.setWidget(content)
        root_layout.addWidget(scroll)

        # --- Кошельки (приватные ключи / мнемоники) ---
        self._lbl_files_header = QLabel()
        layout.addWidget(self._lbl_files_header)

        self._wallet_drop = DropZone(
            label=tr("wallets_label"),
            placeholder="private_keys.txt",
        )
        self._wallet_drop.file_dropped.connect(self._on_wallet_file)
        layout.addWidget(self._wallet_drop)

        # --- Прокси ---
        self._proxy_drop = DropZone(
            label=tr("proxy_label"),
            placeholder="proxy.txt",
        )
        self._proxy_drop.file_dropped.connect(self._on_proxy_file)
        layout.addWidget(self._proxy_drop)

        # --- Фильтры ---
        self._grp_filters = QGroupBox()
        fg = QVBoxLayout(self._grp_filters)

        min_row = QHBoxLayout()
        self._lbl_min_value = QLabel()
        min_row.addWidget(self._lbl_min_value)
        self._min_token_usd = QDoubleSpinBox()
        self._min_token_usd.setRange(0.01, 10000.0)
        self._min_token_usd.setValue(0.03)
        self._min_token_usd.setDecimals(2)
        self._min_token_usd.setFixedWidth(104)
        self._min_token_usd.setToolTip(tr("min_value_tooltip"))
        min_row.addWidget(self._min_token_usd)
        min_row.addStretch()
        fg.addLayout(min_row)

        excl_row = QHBoxLayout()
        self._lbl_exclude_chains = QLabel()
        excl_row.addWidget(self._lbl_exclude_chains)
        self._excluded_chains = ChainPickerWidget(parent=self)
        excl_row.addWidget(self._excluded_chains)
        fg.addLayout(excl_row)

        slippage_row = QHBoxLayout()
        self._lbl_slippage = QLabel()
        slippage_row.addWidget(self._lbl_slippage)
        self._slippage = QDoubleSpinBox()
        self._slippage.setRange(0.1, 50.0)
        self._slippage.setValue(3.0)
        self._slippage.setDecimals(1)
        self._slippage.setFixedWidth(104)
        self._slippage.setToolTip(tr("slippage_tooltip"))
        slippage_row.addWidget(self._slippage)
        slippage_row.addStretch()
        fg.addLayout(slippage_row)

        layout.addWidget(self._grp_filters)

        # --- Бридж ---
        self._grp_bridge = QGroupBox()
        bg = QVBoxLayout(self._grp_bridge)

        target_row = QHBoxLayout()
        self._lbl_target_chains = QLabel()
        target_row.addWidget(self._lbl_target_chains)
        self._target_chains = ChainPickerWidget(parent=self)
        target_row.addWidget(self._target_chains)
        bg.addLayout(target_row)

        min_bridge_row = QHBoxLayout()
        self._lbl_min_bridge = QLabel()
        min_bridge_row.addWidget(self._lbl_min_bridge)
        self._min_bridge_usd = QDoubleSpinBox()
        self._min_bridge_usd.setRange(0.0, 100000.0)
        self._min_bridge_usd.setValue(0.03)
        self._min_bridge_usd.setDecimals(2)
        self._min_bridge_usd.setFixedWidth(104)
        self._min_bridge_usd.setToolTip(tr("min_bridge_tooltip"))
        min_bridge_row.addWidget(self._min_bridge_usd)
        min_bridge_row.addStretch()
        bg.addLayout(min_bridge_row)

        layout.addWidget(self._grp_bridge)

        # --- Задержки ---
        self._grp_delays = QGroupBox()
        dg = QVBoxLayout(self._grp_delays)

        delay_row = QHBoxLayout()
        self._lbl_delay_between_wallets = QLabel()
        delay_row.addWidget(self._lbl_delay_between_wallets)
        self._delay_min = QSpinBox()
        self._delay_min.setRange(0, 3600)
        self._delay_min.setValue(60)
        self._delay_min.setFixedWidth(86)
        self._delay_min.setToolTip(tr("delay_min_tooltip"))
        delay_row.addWidget(self._delay_min)
        delay_row.addWidget(QLabel("—"))
        self._delay_max = QSpinBox()
        self._delay_max.setRange(0, 3600)
        self._delay_max.setValue(180)
        self._delay_max.setFixedWidth(86)
        self._delay_max.setToolTip(tr("delay_max_tooltip"))
        delay_row.addWidget(self._delay_max)
        delay_row.addStretch()
        dg.addLayout(delay_row)

        layout.addWidget(self._grp_delays)

        # --- Субаккаунт биржи ---
        self._grp_exchange = QGroupBox()
        eg = QVBoxLayout(self._grp_exchange)

        self._send_to_exchange = QCheckBox()
        self._send_to_exchange.toggled.connect(self._on_exchange_toggled)
        eg.addWidget(self._send_to_exchange)

        self._sub_drop = DropZone(
            label=tr("subaccounts_label"),
            placeholder="subaccounts.txt",
        )
        self._sub_drop.file_dropped.connect(self._on_subaccount_file)
        self._sub_drop.setEnabled(False)
        eg.addWidget(self._sub_drop)

        delay_bridge_row = QHBoxLayout()
        self._lbl_after_bridge_delay = QLabel()
        delay_bridge_row.addWidget(self._lbl_after_bridge_delay)
        self._delay_after_bridge = QSpinBox()
        self._delay_after_bridge.setRange(0, 3600)
        self._delay_after_bridge.setValue(60)
        self._delay_after_bridge.setFixedWidth(86)
        self._delay_after_bridge.setEnabled(False)
        self._delay_after_bridge.setToolTip(tr("after_bridge_delay_tooltip"))
        delay_bridge_row.addWidget(self._delay_after_bridge)
        delay_bridge_row.addStretch()
        eg.addLayout(delay_bridge_row)

        layout.addWidget(self._grp_exchange)

        # --- Экспорт ---
        self._export_group = QGroupBox()
        self._grp_export = self._export_group
        self._export_group.setEnabled(False)
        exp_layout = QHBoxLayout(self._export_group)
        self._btn_csv  = QPushButton()
        self._btn_json = QPushButton()
        self._btn_xlsx = QPushButton()
        self._btn_csv.clicked.connect(lambda: self._on_export("csv"))
        self._btn_json.clicked.connect(lambda: self._on_export("json"))
        self._btn_xlsx.clicked.connect(lambda: self._on_export("xlsx"))
        exp_layout.addWidget(self._btn_csv)
        exp_layout.addWidget(self._btn_json)
        exp_layout.addWidget(self._btn_xlsx)
        layout.addWidget(self._export_group)

        layout.addStretch()

        i18n.language_changed.connect(self.retranslate_ui)
        self.retranslate_ui()

    def retranslate_ui(self) -> None:
        self._lbl_files_header.setText(tr("files_header"))
        self._wallet_drop.set_label(tr("wallets_label"))
        self._proxy_drop.set_label(tr("proxy_label"))
        self._grp_filters.setTitle(tr("filters_group"))
        self._lbl_min_value.setText(tr("min_value_label"))
        self._lbl_exclude_chains.setText(tr("exclude_chains_label"))
        self._lbl_slippage.setText(tr("slippage_label"))
        self._grp_bridge.setTitle(tr("bridge_group"))
        self._lbl_target_chains.setText(tr("target_chains_label"))
        self._lbl_min_bridge.setText(tr("min_value_label"))
        self._grp_delays.setTitle(tr("delays_group"))
        self._lbl_delay_between_wallets.setText(tr("delay_between_wallets_label"))
        self._grp_exchange.setTitle(tr("exchange_group"))
        self._send_to_exchange.setText(tr("send_to_exchange_checkbox"))
        self._sub_drop.set_label(tr("subaccounts_label"))
        self._lbl_after_bridge_delay.setText(tr("after_bridge_delay_label"))
        self._grp_export.setTitle(tr("export_group"))
        self._btn_csv.setText(tr("save_csv"))
        self._btn_json.setText(tr("save_json"))
        self._btn_xlsx.setText(tr("save_xlsx"))

    # --- Загрузка файлов ---

    def _on_wallet_file(self, path: str) -> None:
        try:
            parsed = parse_wallets(path)
            self._wallets = [w for w in parsed if w["type"] in ("private_key", "mnemonic")]
            self._wallet_drop.set_status(tr("n_wallets_loaded").format(n=len(self._wallets)))
            self._check_wallet_subaccount_match()
        except Exception as e:
            self._wallets = []
            self._wallet_drop.set_status(tr("error_fmt").format(e=e))

    def _on_proxy_file(self, path: str) -> None:
        try:
            self._proxies = parse_proxies(path)
            self._proxy_drop.set_status(tr("n_proxies_loaded").format(n=len(self._proxies)))
        except Exception as e:
            self._proxies = []
            self._proxy_drop.set_status(tr("error_fmt").format(e=e))

    def _on_subaccount_file(self, path: str) -> None:
        import re
        try:
            lines = parse_lines(path)
            invalid = [l for l in lines if not re.fullmatch(r"0x[0-9a-fA-F]{40}", l)]
            if invalid:
                self._subaccounts = []
                self._sub_drop.set_status(
                    tr("error_fmt").format(e=f"Invalid EVM addresses: {', '.join(invalid[:3])}{'…' if len(invalid) > 3 else ''}")
                )
                return
            self._subaccounts = lines
            self._check_wallet_subaccount_match()
        except Exception as e:
            self._subaccounts = []
            self._sub_drop.set_status(tr("error_fmt").format(e=e))

    def _check_wallet_subaccount_match(self) -> None:
        """Показывает предупреждение если кол-во кошельков != субаккаунтов."""
        if not self._send_to_exchange.isChecked():
            return
        w = len(self._wallets)
        s = len(self._subaccounts)
        if w == 0 or s == 0:
            self._sub_drop.set_status("")
            return
        if w != s:
            self._sub_drop.set_status(tr("wallets_subaccounts_mismatch").format(w=w, s=s))
        else:
            self._sub_drop.set_status(tr("n_subaccounts_loaded").format(s=s))

    def _on_exchange_toggled(self, checked: bool) -> None:
        self._sub_drop.setEnabled(checked)
        self._delay_after_bridge.setEnabled(checked)
        self._check_wallet_subaccount_match()

    # --- Публичные методы ---

    def get_wallets(self) -> list[dict]:
        return self._wallets

    def get_proxies(self) -> list[ProxyConfig]:
        return self._proxies

    def get_subaccounts(self) -> list[str]:
        return self._subaccounts

    def get_settings(self) -> CollectorSettings:
        excluded = self._excluded_chains.get_selected()
        target = self._target_chains.get_selected()
        return CollectorSettings(
            min_token_usd=self._min_token_usd.value(),
            min_bridge_usd=self._min_bridge_usd.value(),
            excluded_chains=excluded,
            target_chains=target,
            slippage=self._slippage.value() / 100.0,
            delay_min=self._delay_min.value(),
            delay_max=self._delay_max.value(),
            send_to_exchange=self._send_to_exchange.isChecked(),
            delay_after_bridge=self._delay_after_bridge.value(),
        )

    def on_run_complete(self, results: list, details: dict) -> None:
        """Слот: активирует экспорт после завершения run()."""
        self._export_results = results
        self._export_details = details
        self._export_group.setEnabled(True)

    def _on_export(self, fmt: str) -> None:
        if not self._export_results:
            return
        extensions = {
            "csv":  (tr("save_csv"),  "collector_results.csv",  "CSV (*.csv)"),
            "json": (tr("save_json"), "collector_results.json", "JSON (*.json)"),
            "xlsx": (tr("save_xlsx"), "collector_results.xlsx", "Excel (*.xlsx)"),
        }
        title, default_name, file_filter = extensions[fmt]
        path, _ = QFileDialog.getSaveFileName(self, title, default_name, file_filter)
        if not path:
            return
        import json, csv

        # Build summary rows (bridge_ops excluded — shown in separate Bridges sheet/section)
        summary_rows = []
        bridge_rows = []
        for r in self._export_results:
            data = dict(r.data)
            ops = data.pop("bridge_ops", []) or []
            bridge_count = len(ops)
            summary_rows.append({"item": r.item, "status": r.status.value,
                                  "bridges": bridge_count, **data})
            for op in ops:
                bridge_rows.append({
                    "address": r.item,
                    "src_chain": op.get("src", ""),
                    "tgt_chain": op.get("tgt", ""),
                    "tx_hash": op.get("tx", ""),
                    "status": op.get("status", ""),
                    "sent_usd": op.get("usd", 0),
                })

        total_usd = sum(r.data.get("total_collected_usd", 0) or 0 for r in self._export_results)

        if fmt == "json":
            with open(path, "w", encoding="utf-8") as f:
                json.dump({
                    "total_collected_usd": round(total_usd, 2),
                    "summary": summary_rows,
                    "bridges": bridge_rows,
                }, f, ensure_ascii=False, indent=2)

        elif fmt == "csv":
            if summary_rows:
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.DictWriter(f, fieldnames=summary_rows[0].keys())
                    writer.writeheader()
                    writer.writerows(summary_rows)
                    writer.writerow({})
                    writer.writerow({"item": "TOTAL", "total_collected_usd": round(total_usd, 2)})
                    if bridge_rows:
                        writer.writerow({})
                        writer.writerow({"item": "=== BRIDGES ==="})
                with open(path, "a", newline="", encoding="utf-8-sig") as f:
                    if bridge_rows:
                        writer = csv.DictWriter(f, fieldnames=bridge_rows[0].keys())
                        writer.writeheader()
                        writer.writerows(bridge_rows)

        elif fmt == "xlsx":
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                import subprocess, sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()

            # ── Sheet 1: Summary ──────────────────────────────────────────
            ws1 = wb.active
            ws1.title = "Summary"
            if summary_rows:
                header = list(summary_rows[0].keys())
                ws1.append(header)
                for cell in ws1[1]:
                    cell.font = Font(bold=True)
                for row in summary_rows:
                    ws1.append(list(row.values()))
                ws1.append([])
                total_row = ["TOTAL"] + [""] * (len(header) - 1)
                if "total_collected_usd" in header:
                    total_row[header.index("total_collected_usd")] = round(total_usd, 2)
                ws1.append(total_row)
                for cell in ws1[ws1.max_row]:
                    cell.font = Font(bold=True)

            # ── Sheet 2: Bridges ─────────────────────────────────────────
            ws2 = wb.create_sheet("Bridges")
            if bridge_rows:
                b_header = list(bridge_rows[0].keys())
                ws2.append(b_header)
                for cell in ws2[1]:
                    cell.font = Font(bold=True)
                for row in bridge_rows:
                    ws2.append(list(row.values()))

            try:
                wb.save(path)
            except PermissionError:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(
                    self, tr("save_error_title"),
                    tr("save_error_msg").format(path=path),
                )
