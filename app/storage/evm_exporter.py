from __future__ import annotations
import csv
import json
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from app.core.models import Result


@dataclass
class EvmExportConfig:
    summary: bool = True
    tokens: bool = True
    token_filter: str | None = None


def _parse_filter(f: str | None) -> tuple[str, str] | None:
    """
    Разбирает фильтр "name:chain" по первому ':'.
    Возвращает (name_part, chain_part) или None если фильтр отключён.
    None/"" → None. Строка без ':' → None (фильтр отключён).
    """
    if not f:
        return None
    if ":" not in f:
        return None  # без ':' фильтр не работает
    idx = f.index(":")
    return f[:idx].strip(), f[idx + 1:].strip()


def _match_token(t: dict, flt: tuple[str, str] | None) -> bool:
    if flt is None:
        return True
    name_part, chain_part = flt
    if name_part and (t.get("symbol") or "").upper() != name_part.upper():
        return False
    if chain_part and (t.get("chain") or "").lower() != chain_part.lower():
        return False
    return True


class EvmExporter:
    """Строит и экспортирует данные EVM Balance по конфигурации."""

    def build(
        self,
        results: list[Result],
        details: dict[str, dict],
        config: EvmExportConfig,
    ) -> dict[str, list[dict]]:
        tok_flt = _parse_filter(config.token_filter)
        data: dict[str, list[dict]] = {"summary": [], "tokens": []}

        for r in results:
            if config.summary:
                data["summary"].append({
                    "address":    r.item,
                    "total_usd":  r.data.get("total_usd", ""),
                    "tokens_usd": r.data.get("tokens_usd", ""),
                    "tokens":     r.data.get("tokens", ""),
                    "chains":     r.data.get("chains", ""),
                    "status":     r.status.value,
                })

            if r.status.value != "ok":
                continue

            detail = details.get(r.item, {})
            if config.tokens:
                for t in detail.get("tokens_data", []):
                    if _match_token(t, tok_flt):
                        data["tokens"].append({
                            "address": r.item,
                            "symbol":  t.get("symbol", ""),
                            "chain":   t.get("chain", ""),
                            "amount":  t.get("amount", 0),
                            "price":   t.get("price", 0),
                            "value":   t.get("value", 0),
                        })

        return data

    def export(
        self,
        results: list[Result],
        details: dict[str, dict],
        config: EvmExportConfig,
        path: str,
        fmt: str,
    ) -> None:
        data = self.build(results, details, config)
        if fmt == "csv":
            self._export_csv(data, config, path)
        elif fmt == "json":
            self._export_json(data, path)
        else:
            self._export_xlsx(data, config, path)

    # --- CSV ---

    _SECTION_COLUMNS = {
        "summary": ["address", "total_usd", "tokens_usd", "tokens", "chains", "status"],
        "tokens":  ["address", "symbol", "chain", "amount", "price", "value"],
    }
    _SECTION_LABELS = {
        "summary": "Summary",
        "tokens":  "Tokens",
    }
    _SECTION_FLAGS = {
        "summary": "summary",
        "tokens":  "tokens",
    }

    def _active_sections(self, config: EvmExportConfig) -> list[str]:
        return [k for k in ("summary", "tokens") if getattr(config, self._SECTION_FLAGS[k])]

    def _export_csv(self, data: dict, config: EvmExportConfig, path: str) -> None:
        active = self._active_sections(config)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for i, section in enumerate(active):
                if i > 0:
                    writer.writerow([])  # пустая строка-разделитель
                writer.writerow([f"=== {self._SECTION_LABELS[section]} ==="])
                cols = self._SECTION_COLUMNS[section]
                writer.writerow(cols)
                for row in data[section]:
                    writer.writerow([row.get(c, "") for c in cols])

    # --- JSON ---

    def _export_json(self, data: dict, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # --- XLSX ---

    def _export_xlsx(self, data: dict, config: EvmExportConfig, path: str) -> None:
        active = self._active_sections(config)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # удаляем дефолтный лист

        if not active:
            ws = wb.create_sheet("Summary")
            ws.append(self._SECTION_COLUMNS["summary"])
        else:
            sheet_names = {"summary": "Summary", "tokens": "Tokens"}
            for section in active:
                ws = wb.create_sheet(sheet_names[section])
                cols = self._SECTION_COLUMNS[section]
                ws.append(cols)
                for row in data[section]:
                    ws.append([row.get(c, "") for c in cols])

        wb.save(path)
