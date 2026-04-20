from __future__ import annotations
import csv
import json
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from app.core.models import Result, ResultStatus


@dataclass
class SvmExportConfig:
    summary: bool = True
    tokens: bool = True
    token_filter: str | None = None  # None = all; plain symbol e.g. "USDC"


class SvmExporter:
    _SECTION_COLUMNS = {
        "summary": ["address", "sol_balance", "sol_usd", "total_usd", "tokens", "status"],
        "tokens":  ["address", "symbol", "mint", "amount", "price", "value"],
    }
    _SECTION_LABELS = {"summary": "Summary", "tokens": "Tokens"}

    def build(
        self,
        results: list[Result],
        details: dict[str, dict],
        config: SvmExportConfig,
    ) -> dict[str, list[dict]]:
        data: dict[str, list[dict]] = {"summary": [], "tokens": []}

        for r in results:
            if config.summary:
                data["summary"].append({
                    "address":     r.item,
                    "sol_balance": r.data.get("sol_balance", ""),
                    "sol_usd":     r.data.get("sol_usd", ""),
                    "total_usd":   r.data.get("total_usd", ""),
                    "tokens":      r.data.get("tokens", ""),
                    "status":      r.status.value,
                })

            if r.status != ResultStatus.OK or not config.tokens:
                continue

            detail = details.get(r.item, {})
            for token in detail.get("tokens_data", []):
                if config.token_filter is not None:
                    if token.get("symbol", "").upper() != config.token_filter.upper():
                        continue
                data["tokens"].append({
                    "address": r.item,
                    "symbol":  token.get("symbol", ""),
                    "mint":    token.get("mint", ""),
                    "amount":  token.get("amount", 0),
                    "price":   token.get("price", 0),
                    "value":   token.get("value", 0),
                })

        return data

    def export(
        self,
        results: list[Result],
        details: dict[str, dict],
        config: SvmExportConfig,
        path: str,
        fmt: str,
    ) -> None:
        data = self.build(results, details, config)
        if fmt == "csv":
            self._export_csv(data, config, path)
        elif fmt == "json":
            self._export_json(data, path)
        else:
            self._export_xlsx(data, config, path)

    def _active_sections(self, config: SvmExportConfig) -> list[str]:
        return [k for k in ("summary", "tokens") if getattr(config, k)]

    def _export_csv(self, data: dict, config: SvmExportConfig, path: str) -> None:
        active = self._active_sections(config)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for i, section in enumerate(active):
                if i > 0:
                    writer.writerow([])
                writer.writerow([f"=== {self._SECTION_LABELS[section]} ==="])
                cols = self._SECTION_COLUMNS[section]
                writer.writerow(cols)
                for row in data[section]:
                    writer.writerow([row.get(c, "") for c in cols])

    def _export_json(self, data: dict, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _export_xlsx(self, data: dict, config: SvmExportConfig, path: str) -> None:
        active = self._active_sections(config)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        if not active:
            ws = wb.create_sheet("Summary")
            ws.append(self._SECTION_COLUMNS["summary"])
        else:
            sheet_names = {"summary": "Summary", "tokens": "Tokens"}
            for section in active:
                ws = wb.create_sheet(sheet_names[section])
                cols = self._SECTION_COLUMNS[section]
                ws.append(cols)
                for row in data[section]:
                    ws.append([row.get(c, "") for c in cols])

        wb.save(path)
